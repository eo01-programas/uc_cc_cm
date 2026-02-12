import os
import re
import sys
import time
import shutil
import platform
import subprocess
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from PIL import Image, ImageTk
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ==========================
#  Utilidades de rutas/recursos
# ==========================

def app_dir() -> Path:
    """Devuelve el directorio base de la app (soporta PyInstaller)."""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)  # type: ignore[attr-defined]
    return Path(__file__).resolve().parent

BASE_DIR = app_dir()
DEFAULT_ASSET_DIR = BASE_DIR / "assets"


def locate_asset(base_name: str, exts: list[str], extra_dirs: Optional[list[Path]] = None) -> str:
    """Busca un archivo por nombre base + extensión en directorios razonables."""
    search_dirs: list[Path] = [BASE_DIR / "assets", BASE_DIR, Path.cwd() / "assets", Path.cwd()]
    env_dir = os.environ.get("APP_ASSET_DIR")
    if env_dir and Path(env_dir).exists():
        search_dirs.append(Path(env_dir))
    if DEFAULT_ASSET_DIR.exists():
        search_dirs.append(DEFAULT_ASSET_DIR)
    if extra_dirs:
        # También buscar en subcarpetas assets de los directorios extra
        for d in extra_dirs:
            search_dirs.extend([d / "assets", d])

    for d in search_dirs:
        for ext in exts:
            p = d / f"{base_name}{ext}"
            if p.exists():
                return str(p)
    return ""

# ==========================
#  Normalización de tallas
# ==========================

SIZE_MAP: dict[str, str] = {
    "XXS": "XXS",
    "XS": "XS",
    "S": "S",
    "M": "M",
    "L": "L",
    "XL": "XL",
    "XXL": "2XL",
    "2XL": "2XL",
    "XXXL": "3XL",
    "3XL": "3XL",
    "XSS": "XS",
    "SMALL": "S",
    "MEDIUM": "M",
    "LARGE": "L",
    "EXTRA SMALL": "XS",
    "EXTRA LARGE": "XL",
    "EXTRA EXTRA LARGE": "2XL",
    "EXTRA EXTRA EXTRA LARGE": "3XL",
    "CHICO": "S",
    "MEDIANO": "M",
    "GRANDE": "L",
}
SIZE_ORDER = ["XXS", "XS", "S", "M", "L", "XL", "2XL", "3XL"]
SIZE_CANONICAL = set(SIZE_ORDER)


def norm_size(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    su = str(s).strip().upper()
    return SIZE_MAP.get(su, su)

# ==========================
#  Estado de aplicación
# ==========================

@dataclass
class AppState:
    img1_path: str = ""
    template_excel_path: str = ""

# ==========================
#  Ventana de procesamiento (thread-safe)
# ==========================

class ProcessingWindow:
    def __init__(self, parent: tk.Tk):
        self.parent = parent
        self.window = tk.Toplevel(parent)
        self.window.title("Procesando…")
        self.window.geometry("420x200")
        self.window.resizable(False, False)
        self.window.transient(parent)
        self.window.grab_set()
        self.window.configure(bg="#f7f7f7")

        main_frame = tk.Frame(self.window, bg="#f7f7f7", padx=30, pady=24)
        main_frame.pack(expand=True, fill="both")

        title_label = tk.Label(
            main_frame,
            text="Procesando Documentos",
            font=("Segoe UI", 14, "bold"),
            bg="#f7f7f7",
            fg="#2c3e50",
        )
        title_label.pack(pady=(0, 16))

        self.progress = ttk.Progressbar(
            main_frame,
            length=320,
            mode="indeterminate",
        )
        self.progress.pack(pady=(0, 12))

        self.status_label = tk.Label(
            main_frame,
            text="Por favor, espere mientras procesamos los archivos…",
            font=("Segoe UI", 10),
            bg="#f7f7f7",
            fg="#7f8c8d",
            wraplength=360,
            justify="center",
        )
        self.status_label.pack()

        warning_label = tk.Label(
            main_frame,
            text="⚠️ No cierre esta ventana ni haga clic repetidamente",
            font=("Segoe UI", 9),
            bg="#f7f7f7",
            fg="#e74c3c",
        )
        warning_label.pack(pady=(16, 0))

        self._center_over_parent()
        self.progress.start(10)
        self.window.protocol("WM_DELETE_WINDOW", lambda: None)

    def _center_over_parent(self) -> None:
        self.window.update_idletasks()
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        window_width = self.window.winfo_width()
        window_height = self.window.winfo_height()
        x = parent_x + (parent_width // 2) - (window_width // 2)
        y = parent_y + (parent_height // 2) - (window_height // 2)
        self.window.geometry(f"+{x}+{y}")

    # Métodos seguros para llamar desde hilos secundarios
    def update_status(self, text: str) -> None:
        self.status_label.after(0, lambda: self.status_label.config(text=text))

    def close(self) -> None:
        def _close():
            self.progress.stop()
            self.window.grab_release()
            self.window.destroy()
        self.window.after(0, _close)

# ==========================
#  UI: Preview / cambio imagen
# ==========================

def mostrar_preview(path: str, label: tk.Label) -> None:
    try:
        if not path or not os.path.exists(path):
            raise FileNotFoundError(path or "(vacío)")
        img = Image.open(path)
        img.thumbnail((120, 120))
        photo = ImageTk.PhotoImage(img)
        label.config(image=photo, text="")
        label.image = photo  # evitar GC
    except Exception:
        label.config(text="(sin imagen)", image="")
        label.image = None

# ==========================
#  Ventana: ingreso Case QTY por estilo
# ==========================

def pedir_case_qty_por_estilo(root: tk.Tk, estilos: Iterable[str], valores_default: Optional[dict[str, int]] = None) -> Optional[dict[str, int]]:
    top = tk.Toplevel(root)
    top.title("Case QTY por estilo")
    top.grab_set()
    top.geometry("540x460")

    tk.Label(
        top,
        text="Revisa/edita Case QTY por estilo (valores del Excel mostrados):",
        anchor="w",
        justify="left",
        font=("Segoe UI", 10)
    ).pack(padx=10, pady=(10, 6), fill="x")

    fill_frame = tk.Frame(top)
    fill_frame.pack(fill="x", padx=10, pady=4)
    tk.Label(fill_frame, text="Rellenar todo con:", font=("Segoe UI", 9)).pack(side="left")
    entry_all = tk.Entry(fill_frame, width=8)
    entry_all.pack(side="left", padx=6)
    entries: dict[str, tk.Entry] = {}

    def aplicar_todo() -> None:
        val = entry_all.get().strip()
        for e in entries.values():
            e.delete(0, tk.END)
            e.insert(0, val)

    ttk.Button(fill_frame, text="Aplicar", command=aplicar_todo).pack(side="left")

    canvas = tk.Canvas(top, borderwidth=0)
    frame = tk.Frame(canvas)
    vsb = tk.Scrollbar(top, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vsb.set)
    vsb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True, padx=10, pady=6)
    canvas.create_window((0, 0), window=frame, anchor="nw")

    for i, est in enumerate(sorted(estilos)):
        tk.Label(frame, text=est, anchor="w", width=42, font=("Segoe UI", 9)).grid(row=i, column=0, sticky="w", pady=2)
        e = tk.Entry(frame, width=10)
        e.grid(row=i, column=1, sticky="w", pady=2, padx=(6, 0))
        if valores_default and est in valores_default:
            e.insert(0, str(valores_default[est]))
        entries[est] = e

    def on_configure(_):
        canvas.configure(scrollregion=canvas.bbox("all"))

    frame.bind("<Configure>", on_configure)

    result: Optional[dict[str, int]] = {}

    def aceptar() -> None:
        nonlocal result
        try:
            res: dict[str, int] = {}
            for est, e in entries.items():
                s = e.get().strip()
                qty = 0 if s == "" else int(float(s))
                if qty < 0:
                    raise ValueError
                res[est] = qty
            result = res
            top.destroy()
        except Exception:
            messagebox.showerror("Valor inválido", "Case QTY debe ser un número entero por estilo.")

    def cancelar() -> None:
        nonlocal result
        result = None
        top.destroy()

    btns = tk.Frame(top)
    btns.pack(pady=8)
    ttk.Button(btns, text="Aceptar", width=14, command=aceptar).pack(side="left", padx=6)
    ttk.Button(btns, text="Cancelar", width=14, command=cancelar).pack(side="left", padx=6)

    top.wait_window()
    return result

# ==========================
#  PDF: detección / extracción
# ==========================

def detectar_formato(pdf_path: str) -> str:
    try:
        with pdfplumber.open(pdf_path) as doc:
            text = (doc.pages[0].extract_text() or "")
            if "Division|" in text:
                return "Barras"
            if "UPC REPORT" in text:
                return "Matricial"
    except Exception:
        pass
    return "Desconocido"


def extract_data_barras(pdf_path: str) -> list[dict[str, str]]:
    data: list[dict[str, str]] = []
    with pdfplumber.open(pdf_path) as doc:
        full_text = "\n".join([page.extract_text() or "" for page in doc.pages])
    lines = [ln.strip() for ln in full_text.split("\n") if ln.strip()]
    for line in lines:
        if "Division|" in line and "Style|" in line and "UPC|" in line:
            continue
        if "|" not in line:
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 8:
            continue
        _, style, upc, _, color_code, color_name, _, size = parts[:8]
        upc_clean = re.sub(r"\D", "", upc)
        if not upc_clean or not upc_clean.isdigit():
            continue
        row = {
            "STYLE": str(style).strip().upper(),
            "COLOR CODE": str(color_code).strip().upper(),
            "COLOR NAME": str(color_name).strip().upper(),
            "SIZE": str(size).strip().upper(),
            "UPC CODE": upc_clean,
            "STYLE COLOR": f"{str(style).strip().upper()} {str(color_code).strip().upper()}",
        }
        data.append(row)
    return data


def extract_data_matricial(pdf_path: str) -> list[dict[str, str]]:
    registros: list[dict[str, str]] = []
    style_actual: Optional[str] = None
    tallas_actuales: list[str] = []

    style_re = re.compile(r"^(TP\d+[A-Z]?)\b")
    size_token = re.compile(r"\*+\s*([A-Z0-9/]+)\s*\*+")
    color_line = re.compile(r"^([A-Z0-9]{3,4})\s+([A-Z0-9/ .\-]+?)(?:\s+((?:\d{11,14}\s+)*\d{11,14}))?\s*$")
    numbers_only = re.compile(r"^(?:\d{11,14}\s+)*\d{11,14}$")

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            raw_lines = [ln.rstrip() for ln in text.split("\n")]
            i = 0
            while i < len(raw_lines):
                line = raw_lines[i].strip()
                if not line or line.startswith("-") or line.startswith("*"):
                    i += 1
                    continue

                m_style = style_re.match(line)
                if m_style:
                    style_actual = m_style.group(1).upper()
                    tallas_actuales = size_token.findall(line)
                    j = i + 1
                    while j < len(raw_lines):
                        nxt = raw_lines[j].strip()
                        if not nxt:
                            break
                        if style_re.match(nxt) or color_line.match(nxt):
                            break
                        extra_sizes = size_token.findall(nxt)
                        if not extra_sizes:
                            break
                        tallas_actuales.extend(extra_sizes)
                        j += 1
                    i = j
                    continue

                m_color = color_line.match(line)
                if m_color and style_actual and tallas_actuales:
                    color_code = m_color.group(1).upper()
                    color_name = m_color.group(2).strip().upper()
                    upcs: list[str] = []
                    if m_color.group(3):
                        upcs.extend(m_color.group(3).split())
                    k = i + 1
                    while k < len(raw_lines):
                        nxt = raw_lines[k].strip()
                        if numbers_only.match(nxt):
                            upcs.extend(nxt.split())
                            k += 1
                            continue
                        if color_line.match(nxt) or style_re.match(nxt):
                            break
                        if not nxt:
                            break
                        break
                    n = min(len(tallas_actuales), len(upcs))
                    for idx in range(n):
                        registros.append(
                            {
                                "STYLE": style_actual,
                                "COLOR CODE": color_code,
                                "COLOR NAME": color_name,
                                "SIZE": tallas_actuales[idx].upper(),
                                "UPC CODE": upcs[idx],
                                "STYLE COLOR": f"{style_actual} {color_code}",
                            }
                        )
                    i = k
                    continue

                i += 1
    return registros


def extract_data_from_pdf(pdf: str) -> list[dict[str, str]]:
    tipo = detectar_formato(pdf)
    if tipo == "Barras":
        return extract_data_barras(pdf)
    rows = extract_data_matricial(pdf)
    if not rows and tipo == "Desconocido":
        rows = extract_data_barras(pdf)
    return rows

# ==========================
#  EXCEL: preparar datos (solo USA) + columnas extra
# ==========================

def _rename_canonical(df: pd.DataFrame) -> pd.DataFrame:
    col_map_uc = {
        "STYLE": "NOMBRE ESTILO",
        "ESTILOS": "NOMBRE ESTILO",
        "NOMBRE ESTILO": "NOMBRE ESTILO",
        "OP": "PEDIDO PRODUCCION COFACO",
        "PEDIDO PRODUCCION COFACO": "PEDIDO PRODUCCION COFACO",
        "PROTO": "PROTO COFACO",
        "PROTO COFACO": "PROTO COFACO",
        "DESTINO": "DESTINO",
        "PO#": "PO#",
        "PO": "PO#",
        "PO NO": "PO#",
        "PO NO.": "PO#",
        "DESCRIPCION COLOR": "NOMBRE COLOR",
        "NOMBRE COLOR": "NOMBRE COLOR",
        "COLOR": "COLOR",
        "CARTA": "CARTA",
        "COLR CODE": "COLOR CODE",
        "COLOR CODE": "COLOR CODE",
        "COLUMNA1": "HOJA MARCACION",
        "HOJA DE MARCACIÓN": "HOJA MARCACION",
        "HOJA DE MARCACION": "HOJA MARCACION",
        "TOTAL": "UNITS/TALLA (PEDIDO)",
        "UNITS/TALLA (PEDIDO)": "UNITS/TALLA (PEDIDO)",
        "SKX PO#": "SKX PO#",
        "SKX PO": "SKX PO#",
        "WIP LINE NUMBER": "WIP LINE NUMBER",
        "WIP LINE NUMBER:": "WIP LINE NUMBER",
        "LN": "WIP LINE NUMBER",
        "CASE QTY": "CASE QTY",
        "CASEQTY": "CASE QTY",
    }
    idx = {k.upper(): v for k, v in col_map_uc.items()}
    return df.rename(columns={c: idx.get(str(c).strip().upper(), c) for c in df.columns})


def _read_excel_flexible(excel_path: str) -> pd.DataFrame:
    """Intenta leer el Excel detectando hoja y fila de encabezados automáticamente."""
    def norm_token(value: object) -> str:
        s = str(value).strip().upper()
        s = re.sub(r"[^A-Z0-9#]+", " ", s)
        return re.sub(r"\s+", " ", s).strip()

    header_terms = {
        "STYLE",
        "ESTILOS",
        "OP",
        "PROTO",
        "DESTINO",
        "PO",
        "PO NO",
        "PO NO.",
        "PO#",
        "DESCRIPCION COLOR",
        "COLOR",
        "CARTA",
        "CODE",
        "COLR CODE",
        "COLOR CODE",
    }

    best: Optional[tuple[int, str, int]] = None  # (hits, sheet, header_row)

    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
        for sheet in xls.sheet_names:
            preview = pd.read_excel(excel_path, engine="openpyxl", sheet_name=sheet, header=None, nrows=30)
            for idx, row in preview.iterrows():
                cells = [norm_token(c) for c in row.tolist()]
                hits = {c for c in cells if c in header_terms}
                if "STYLE" in hits or "ESTILOS" in hits:
                    if best is None or len(hits) > best[0]:
                        best = (len(hits), sheet, idx)
    except Exception as e:
        print(f"Error inspeccionando hojas del Excel: {e}")
        best = None

    if best:
        _, sheet, header_row = best
        df = pd.read_excel(excel_path, dtype=str, engine="openpyxl", sheet_name=sheet, header=header_row)
        if df is not None and df.shape[1] > 0:
            df.columns = [str(col).strip() if pd.notna(col) else f"Col_{i}" for i, col in enumerate(df.columns)]
            print(f"Excel detectado: hoja='{sheet}', header={header_row}")
            print(f"Columnas finales: {list(df.columns)}")
            return df

    # Fallback al comportamiento anterior (hoja 0)
    df_preview = pd.read_excel(excel_path, engine="openpyxl", sheet_name=0, header=None, nrows=10)
    best_header_row = 0
    max_text_score = 0
    for i in range(min(5, len(df_preview))):
        row = df_preview.iloc[i]
        text_score = 0
        for val in row:
            if pd.notna(val) and isinstance(val, str):
                val_clean = str(val).strip()
                if len(val_clean) > 2 and not val_clean.replace(".", "").replace(",", "").isdigit():
                    text_score += 1
        if text_score > max_text_score:
            max_text_score = text_score
            best_header_row = i

    try:
        df = pd.read_excel(excel_path, dtype=str, engine="openpyxl", sheet_name=0, header=best_header_row)
        if df is not None and df.shape[1] > 0:
            df.columns = [str(col).strip() if pd.notna(col) else f"Col_{i}" for i, col in enumerate(df.columns)]
            print(f"Columnas finales: {list(df.columns)}")
            return df
    except Exception as e:
        print(f"Error leyendo con header={best_header_row}: {e}")

    try:
        df = pd.read_excel(excel_path, dtype=str, engine="openpyxl", sheet_name=0, header=None)
        df.columns = [f"Col_{i}" for i in range(len(df.columns))]
        print(f"Fallback: usando columnas genéricas {list(df.columns)}")
        return df
    except Exception:
        raise ValueError("No se pudo leer el archivo Excel. Verifica que sea un archivo válido.")


def preparar_excel(df_excel: pd.DataFrame) -> pd.DataFrame:
    df_excel = _rename_canonical(df_excel)

    # Forward-fill para columnas que suelen tener celdas combinadas (merged) en el Excel.
    # Pandas solo lee el valor de la primera celda del rango combinado; las demás quedan NaN.
    # Se hace ffill agrupado por NOMBRE ESTILO (o su alias) para no arrastrar valores entre estilos distintos.
    _ffill_cols = ["CASE QTY", "WIP LINE NUMBER", "TT"]
    _group_col = None
    for _gc in ["NOMBRE ESTILO", "STYLE", "ESTILOS"]:
        if _gc in df_excel.columns:
            _group_col = _gc
            break
    for _fc in _ffill_cols:
        if _fc in df_excel.columns:
            if _group_col:
                df_excel[_fc] = df_excel.groupby(_group_col)[_fc].transform(lambda s: s.ffill())
            else:
                df_excel[_fc] = df_excel[_fc].ffill()

    # Resolver columnas duplicadas (p.ej. múltiples "COLOR CODE")
    if "COLOR CODE" in df_excel.columns and isinstance(df_excel["COLOR CODE"], pd.DataFrame):
        combined = df_excel["COLOR CODE"].bfill(axis=1).iloc[:, 0]
        df_excel = df_excel.drop(columns=["COLOR CODE"])
        df_excel["COLOR CODE"] = combined

    # Ajustes para nuevos formatos (p.ej. COLOR = nombre y COLR CODE = código)
    if "NOMBRE COLOR" not in df_excel.columns:
        if "COLOR" in df_excel.columns:
            df_excel["NOMBRE COLOR"] = df_excel["COLOR"].astype(str)
        elif "CARTA" in df_excel.columns:
            df_excel["NOMBRE COLOR"] = df_excel["CARTA"].astype(str)

    if "COLOR" not in df_excel.columns and "COLOR CODE" in df_excel.columns:
        df_excel["COLOR"] = df_excel["COLOR CODE"].astype(str)

    if "COLOR" in df_excel.columns and "COLOR CODE" in df_excel.columns:
        # Si COLOR es nombre y COLOR CODE es código, priorizar el código en COLOR
        df_excel["COLOR"] = df_excel["COLOR CODE"].astype(str)

    # Soporte: si la plantilla trae la columna "COLUMNA1" (renombrada a "HOJA MARCACION")
    # pero en realidad contiene el WIP Line Number, detectar y copiarla a la columna esperada.
    try:
        if "WIP LINE NUMBER" not in df_excel.columns and "HOJA MARCACION" in df_excel.columns:
            sample = df_excel["HOJA MARCACION"].dropna().astype(str).head(20).tolist()
            if sample:
                import re
                digits_ratio = sum(1 for s in sample if re.search(r"\d", s)) / len(sample)
                # Si una proporción razonable contiene dígitos, asumimos que es WIP
                if digits_ratio >= 0.4:
                    df_excel["WIP LINE NUMBER"] = df_excel["HOJA MARCACION"].astype(str)
                    print('DEBUG: Copiada columna HOJA MARCACION -> WIP LINE NUMBER (detección automática)')
    except Exception:
        pass

    # Normalizar alias comunes de UPC en caso de que Excel tenga otro encabezado
    try:
        if "UPC Barcode" not in df_excel.columns:
            for alt in ["UPC CODE", "UPC", "UPC_BARCODE", "UPC BARCODE"]:
                if alt in df_excel.columns:
                    df_excel["UPC Barcode"] = df_excel[alt].astype(str)
                    print(f'DEBUG: Copiada columna {alt} -> UPC Barcode')
                    break
    except Exception:
        pass

    # Debug: mostrar columnas encontradas
    print(f"Columnas encontradas en Excel: {list(df_excel.columns)}")

    # Si no hay columnas reconocidas, mostrar las primeras filas para diagnóstico
    if all("Unnamed" in str(col) or "Col_" in str(col) for col in df_excel.columns):
        print("El Excel parece no tener encabezados claros. Primeras 3 filas:")
        print(df_excel.head(3).to_string())
        raise ValueError(
            "No se pudieron identificar las columnas necesarias en el Excel.\n\n"
            "El archivo debe tener columnas con nombres como:\n"
            "- NOMBRE ESTILO (o ESTILOS)\n"
            "- DESTINO\n"
            "- PO# (o PO)\n"
            "- NOMBRE COLOR (o DESCRIPCION COLOR)\n"
            "- COLOR\n\n"
            "Verifica que:\n"
            "1. El archivo tenga encabezados en la primera fila\n"
            "2. Los nombres de las columnas estén escritos correctamente\n"
            "3. No haya filas vacías antes de los encabezados"
        )

    requeridas = [
        "NOMBRE ESTILO",
        "DESTINO",
        "PO#",
        "NOMBRE COLOR",
        "COLOR",
    ]
    falt = [c for c in requeridas if c not in df_excel.columns]
    if falt:
        # Mostrar más información para diagnosticar
        available_cols = list(df_excel.columns)
        raise ValueError(f"Faltan columnas requeridas en el Excel: {falt}\n\nColumnas disponibles: {available_cols}\n\nVerifica que el archivo Excel tenga las columnas necesarias o que estén bien escritas.")

    # Add missing optional columns with default values
    if "PEDIDO PRODUCCION COFACO" not in df_excel.columns:
        df_excel["PEDIDO PRODUCCION COFACO"] = ""
    if "PROTO COFACO" not in df_excel.columns:
        df_excel["PROTO COFACO"] = ""

    for c in ["NOMBRE ESTILO", "DESTINO", "NOMBRE COLOR", "COLOR"]:
        df_excel[c] = df_excel[c].astype(str).str.strip().str.upper()

    # ¿Tallas como columnas?
    size_cols: list[str] = []
    for c in df_excel.columns:
        cu = str(c).strip().upper()
        if cu in SIZE_CANONICAL or cu in SIZE_MAP:
            size_cols.append(c)

    id_vars_base = [
        "NOMBRE ESTILO",
        "PEDIDO PRODUCCION COFACO",
        "PROTO COFACO",
        "DESTINO",
        "PO#",
        "NOMBRE COLOR",
        "COLOR",
    ]
    extras: list[str] = []
    for c in ["HOJA MARCACION", "UNITS/TALLA (PEDIDO)", "WIP LINE NUMBER", "CASE QTY", "TT"]:
        if c in df_excel.columns:
            extras.append(c)

    if size_cols:
        canon = {c: SIZE_MAP.get(str(c).strip().upper(), str(c).strip().upper()) for c in size_cols}
        df_long = df_excel.melt(
            id_vars=id_vars_base + extras,
            value_vars=size_cols,
            var_name="SIZE_RAW",
            value_name="QTY",
        )
        df_long["SIZE"] = df_long["SIZE_RAW"].map(lambda x: canon.get(x, str(x).upper()))

        def has_qty(v: object) -> bool:
            if v is None:
                return False
            s = str(v).strip()
            if s == "":
                return False
            try:
                return float(s) > 0
            except Exception:
                return True

        df_long = df_long[df_long["QTY"].apply(has_qty)].copy()
        df_long.rename(columns={"QTY": "QTY POR TALLA"}, inplace=True)
        return df_long.reset_index(drop=True)

    out = df_excel[id_vars_base + extras].drop_duplicates().reset_index(drop=True)
    out["SIZE"] = ""
    out["QTY POR TALLA"] = ""
    return out

# ==========================
#  EXCEL: escritura y fórmulas
# ==========================

def apply_formulas_to_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int, max_col: int, last_data_row: int) -> None:
    """Aplica fórmulas y agrega notas. Mantiene encabezados del template y aplica formato.
    - header_row: fila con encabezados (p. ej. 10)
    - last_data_row: última fila con datos (>= header_row)
    """
    first_data_row = header_row + 1

    headers: dict[str, int] = {}
    for c in range(1, max_col + 1):
        cell_value = ws.cell(row=header_row, column=c).value
        if cell_value:
            headers[str(cell_value)] = c

    col_case = headers.get("Case QTY")
    col_qty_talla = headers.get("QTY POR TALLA")
    col_result = headers.get("QTY DE STICKERS A IMPRIMIR")

    # Fórmula: redondeo hacia arriba por caja + 3
    if col_case and col_qty_talla and col_result and last_data_row >= first_data_row:
        c_case = get_column_letter(col_case)
        c_qty = get_column_letter(col_qty_talla)
        for r in range(first_data_row, last_data_row + 1):
            formula = f'=IFERROR(ROUNDUP({c_qty}{r}/VALUE(SUBSTITUTE({c_case}{r},"Q","")),0)+3,0)'
            cell = ws.cell(row=r, column=col_result)
            cell.value = formula
            cell.number_format = "0"
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Forzar formato de texto para códigos
    if "UPC Barcode" in headers:
        for r in range(first_data_row, last_data_row + 1):
            cell = ws.cell(row=r, column=headers["UPC Barcode"])
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                cell.value = str(cell.value)
            cell.data_type = "s"
            cell.number_format = "@"
    if "Case QTY" in headers:
        for r in range(first_data_row, last_data_row + 1):
            cell = ws.cell(row=r, column=headers["Case QTY"])
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                cell.value = str(cell.value)
            cell.data_type = "s"
            cell.number_format = "@"

    # APLICAR FORMATO SOLO A LOS ENCABEZADOS DE LA TABLA (fila header_row) - NEGRITA Y FONDO AMARILLO
    for c in range(1, max_col + 1):
        header_cell = ws.cell(row=header_row, column=c)
        if header_cell.value:
            header_cell.font = Font(bold=True, color="000000")  # Negro y negrita
            header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo amarillo
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # AGREGAR AJUSTAR TEXTO

    # QUITAR LÍNEAS DE CUADRÍCULA DE LA HOJA
    ws.sheet_view.showGridLines = False

    # Limpiar notas anteriores si existen (plantilla o ejecuciones previas)
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.strip().lower() == "important notes:":
            ws.cell(row=r, column=1).value = None
            ws.cell(row=r, column=1).font = Font()
            ws.cell(row=r, column=1).fill = PatternFill()
            ws.cell(row=r + 1, column=1).value = None

    # NOTAS DINÁMICAS - basadas en la última fila de datos de ESTA hoja específica
    # Dos filas en blanco después de la tabla
    note_header_row = last_data_row + 2

    # APLICAR FORMATO SOLO A LAS NOTAS DINÁMICAS (no a posiciones fijas)
    # "Important notes:" - con negrita y fondo amarillo
    note_header_cell = ws.cell(row=note_header_row, column=1)
    note_header_cell.value = "Important notes:"
    note_header_cell.font = Font(bold=True, color="000000")
    note_header_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Texto de las notas - sin formato especial
    note_text = "SKX PO#:, WIP Line #, UPC Barcode, CASE Qty:  Barcode format= code 128 ( con PO# 73 Barras)"
    note_text_cell = ws.cell(row=note_header_row + 1, column=1)
    note_text_cell.value = note_text
    note_text_cell.font = Font(size=10)
    # NO hacer merge ni wrap_text para que quede en una línea


def find_last_data_row(ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int, max_col: int) -> int:
    """Encuentra la última fila con datos dentro del rango de la tabla."""
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, max_col + 1)):
            last = r
    return max(last, start_row - 1)


# ==========================
#  Imagen en hojas (fix)
# ==========================

def copy_template_header_to_worksheet(template_ws, target_ws, img_path: str):
    """Coloca la imagen seleccionada en la hoja destino.
    - Si la plantilla tiene una imagen, usa su ancla y tamaño.
    - Si no, ancla en A1 y redimensiona si es muy grande.
    - Reemplaza cualquier imagen previa en la hoja destino.
    """
    try:
        if not (img_path and os.path.exists(img_path)):
            return

        from openpyxl.drawing.image import Image as OpenpyxlImage
        img = OpenpyxlImage(img_path)

        # Leer referencia de la imagen de la plantilla (si existe)
        template_images = getattr(template_ws, "_images", [])
        anchor = "A1"
        t_w = t_h = None
        if template_images:
            timg = template_images[0]
            anchor = getattr(timg, "anchor", "A1")
            if hasattr(timg, "width") and hasattr(timg, "height"):
                t_w, t_h = timg.width, timg.height

        # Ajustar tamaño y ancla del nuevo gráfico
        if t_w and t_h:
            img.width = t_w
            img.height = t_h
        else:
            if hasattr(img, "width") and img.width and img.width > 200:
                ratio = 200 / img.width
                img.width = 200
                if hasattr(img, "height") and img.height:
                    img.height = int(img.height * ratio)
        img.anchor = anchor

        # Reemplazar imágenes existentes en destino
        if hasattr(target_ws, "_images") and target_ws._images:
            target_ws._images.clear()

        target_ws.add_image(img)

    except Exception as e:
        print(f"Error colocando imagen: {e}")

# ==========================
#  Sistema
# ==========================

def open_file_and_folder(file_path: str) -> None:
    sistema = platform.system()
    try:
        if sistema == "Windows":
            os.startfile(file_path)  # type: ignore[attr-defined]
            os.startfile(os.path.dirname(file_path))  # type: ignore[attr-defined]
        elif sistema == "Darwin":
            subprocess.Popen(["open", file_path])
            subprocess.Popen(["open", os.path.dirname(file_path)])
        else:
            subprocess.Popen(["xdg-open", file_path])
            subprocess.Popen(["xdg-open", os.path.dirname(file_path)])
    except Exception as e:
        print(f"No se pudo abrir el archivo o carpeta: {e}")

# ==========================
#  Aplicación principal (Tk)
# ==========================

class App:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.state = AppState(
            img1_path=locate_asset("imagen 1", [".png", ".jpg", ".jpeg", ".bmp"]),
            template_excel_path=(
                locate_asset("encabezado", [".xlsx", ".xlsm"]) or
                locate_asset("plantilla", [".xlsx", ".xlsm"])  # fallback
            ),
        )
        self._build_ui()

    # ---------------------- UI ----------------------
    def _build_ui(self) -> None:
        self.root.title("Generador de Reporte Final")
        self.root.geometry("440x280")

        lbl_title = tk.Label(self.root, text="Generador de Reporte Final", font=("Segoe UI", 12, "bold"))
        lbl_title.pack(pady=8)

        frm_imgs = tk.Frame(self.root)
        frm_imgs.pack(pady=4)

        self.lbl_img1 = tk.Label(frm_imgs, text="(sin imagen)")
        self.lbl_img1.grid(row=0, column=0, padx=8, pady=4)
        mostrar_preview(self.state.img1_path, self.lbl_img1)

        ttk.Button(frm_imgs, text="Cargar/Cambiar imagen…", command=self.cambiar_imagen, width=26).grid(row=1, column=0, pady=6)

        ttk.Button(
            self.root,
            text="Procesar (seleccionar PDF y Excel)…",
            command=self.process_all,
            width=36
        ).pack(pady=16)

    def cambiar_imagen(self) -> None:
        file = filedialog.askopenfilename(title="Selecciona la imagen", filetypes=[("Imágenes", "*.png;*.jpg;*.jpeg;*.bmp")])
        if file:
            self.state.img1_path = file
            mostrar_preview(self.state.img1_path, self.lbl_img1)

    # ---------------------- Lógica principal ----------------------
    def _obtener_case_qty_por_estilo(self, estilos: list[str], valores_default: Optional[dict[str, int]] = None) -> dict[str, int]:
        result_holder: list[Optional[dict[str, int]]] = [None]

        def ask() -> None:
            result_holder[0] = pedir_case_qty_por_estilo(self.root, estilos, valores_default)

        self.root.after(0, ask)
        # Espera activa mínima en el hilo secundario hasta que usuario cierre el diálogo
        while result_holder[0] is None:
            time.sleep(0.1)
        mapping = result_holder[0]
        if mapping is None:
            raise Exception("Proceso cancelado por el usuario")
        return mapping

    def process_all(self) -> None:
        pdf_paths = filedialog.askopenfilenames(title="Selecciona PDF(s)", filetypes=[("PDF", "*.pdf")])
        if not pdf_paths:
            return

        excel_path = filedialog.askopenfilename(title="Selecciona el Excel (datos)", filetypes=[("Excel", "*.xlsx;*.xls;*.xlsm")])
        if not excel_path:
            return

        proc = ProcessingWindow(self.root)

        def worker() -> None:
            try:
                proc.update_status("Ubicando recursos…")
                extra_dirs = [Path(pdf_paths[0]).parent, Path(excel_path).parent]
                if not self.state.img1_path or not os.path.exists(self.state.img1_path):
                    self.state.img1_path = locate_asset("imagen 1", [".png", ".jpg", ".jpeg", ".bmp"], extra_dirs)
                if not self.state.template_excel_path or not os.path.exists(self.state.template_excel_path):
                    self.state.template_excel_path = (
                        locate_asset("encabezado", [".xlsx", ".xlsm"], extra_dirs) or
                        locate_asset("plantilla", [".xlsx", ".xlsm"], extra_dirs)
                    )
                if not self.state.template_excel_path or not os.path.exists(self.state.template_excel_path):
                    raise FileNotFoundError(
                        "No se encontró 'encabezado.xlsx/xlsm' (ni 'plantilla.xlsx'). Colócalo junto a los PDFs, al Excel o en 'assets'."
                    )

                # Actualizar preview en el hilo principal
                self.root.after(0, lambda: mostrar_preview(self.state.img1_path, self.lbl_img1))

                proc.update_status("Extrayendo datos de PDFs…")
                all_registros: list[dict[str, str]] = []
                for pdf in pdf_paths:
                    all_registros.extend(extract_data_from_pdf(pdf))

                if not all_registros:
                    raise RuntimeError("No se extrajo información de los PDFs.")

                df_pdfs = pd.DataFrame(all_registros)
                for c in ["STYLE", "COLOR CODE", "COLOR NAME", "SIZE"]:
                    if c in df_pdfs:
                        df_pdfs[c] = df_pdfs[c].astype(str).str.strip().str.upper()
                if "SIZE" in df_pdfs:
                    df_pdfs["SIZE"] = df_pdfs["SIZE"].map(norm_size)

                proc.update_status("Procesando Excel de datos…")
                df_excel_raw = _read_excel_flexible(excel_path)
                df_excel = preparar_excel(df_excel_raw)
                for c in ["NOMBRE ESTILO", "DESTINO", "NOMBRE COLOR", "COLOR", "SIZE"]:
                    if c in df_excel:
                        df_excel[c] = df_excel[c].astype(str).str.strip().str.upper()
                if "SIZE" in df_excel.columns:
                    df_excel["SIZE"] = df_excel["SIZE"].map(norm_size)

                # Filtrar solo filas con DESTINO = USA antes del merge con PDFs
                if "DESTINO" in df_excel.columns:
                    df_excel = df_excel[df_excel["DESTINO"] == "USA"].copy()
                    if df_excel.empty:
                        raise RuntimeError("No se encontraron filas con DESTINO = USA en el Excel.")

                # Merge por nombre/código de color
                if "SIZE" in df_excel.columns:
                    df_name = pd.merge(
                        df_excel,
                        df_pdfs,
                        left_on=["NOMBRE ESTILO", "NOMBRE COLOR", "SIZE"],
                        right_on=["STYLE", "COLOR NAME", "SIZE"],
                        how="inner",
                    )
                    df_code = pd.merge(
                        df_excel,
                        df_pdfs,
                        left_on=["NOMBRE ESTILO", "COLOR", "SIZE"],
                        right_on=["STYLE", "COLOR CODE", "SIZE"],
                        how="inner",
                    )
                else:
                    df_name = pd.merge(
                        df_excel,
                        df_pdfs,
                        left_on=["NOMBRE ESTILO", "NOMBRE COLOR"],
                        right_on=["STYLE", "COLOR NAME"],
                        how="inner",
                    )
                    df_code = pd.merge(
                        df_excel,
                        df_pdfs,
                        left_on=["NOMBRE ESTILO", "COLOR"],
                        right_on=["STYLE", "COLOR CODE"],
                        how="inner",
                    )

                subset_cols = [c for c in [
                    "NOMBRE ESTILO", "NOMBRE COLOR", "COLOR", "DESTINO", "PO#", "SIZE", "UPC CODE"
                ] if c in (df_name.columns.union(df_code.columns))]
                df_merge_all = pd.concat([df_name, df_code], ignore_index=True).drop_duplicates(subset=subset_cols)
                # DEBUG: información para diagnosticar campos vacíos
                try:
                    print("DEBUG: columnas df_excel:", list(df_excel.columns))
                    print("DEBUG: columnas df_pdfs:", list(df_pdfs.columns))
                    print("DEBUG: columnas df_merge_all:", list(df_merge_all.columns))
                    # Revisar WIP Line Number
                    if "WIP Line Number" in df_merge_all.columns:
                        missing_wip = df_merge_all["WIP Line Number"].isna() | (df_merge_all["WIP Line Number"].astype(str).str.strip() == "")
                        print(f"DEBUG: filas totales={len(df_merge_all)}, WIP Line Number vacías={missing_wip.sum()}")
                        if missing_wip.any():
                            print("DEBUG: primeras filas con WIP vacío:")
                            print(df_merge_all.loc[missing_wip].head(5).to_string())
                    else:
                        print('DEBUG: columna "WIP Line Number" no encontrada en df_merge_all')

                    # Revisar UPC Barcode / UPC CODE
                    upc_col = None
                    if "UPC Barcode" in df_merge_all.columns:
                        upc_col = "UPC Barcode"
                    elif "UPC CODE" in df_merge_all.columns:
                        upc_col = "UPC CODE"
                    if upc_col:
                        missing_upc = df_merge_all[upc_col].isna() | (df_merge_all[upc_col].astype(str).str.strip() == "")
                        print(f"DEBUG: UPC column used='{upc_col}', vacías={missing_upc.sum()}")
                        if missing_upc.any():
                            print("DEBUG: primeras filas con UPC vacío:")
                            print(df_merge_all.loc[missing_upc].head(5).to_string())
                    else:
                        print('DEBUG: no se encontró columna UPC en df_merge_all')
                except Exception as _:
                    print('DEBUG: error al imprimir información de diagnóstico')

                if df_merge_all.empty:
                    raise RuntimeError("No hubo intersección entre PDFs y Excel (DESTINO=USA).")

                # Aliases, columnas y reglas
                df_merge_all["PROTO"] = df_merge_all.get("PROTO COFACO", "")
                df_merge_all["OP"] = df_merge_all.get("PEDIDO PRODUCCION COFACO", "")
                df_merge_all["PO(cliente)"] = df_merge_all.get("PO#", "")
                df_merge_all["US Size"] = df_merge_all.get("SIZE", "")
                df_merge_all["QTY POR TALLA"] = pd.to_numeric(df_merge_all.get("QTY POR TALLA", ""), errors="coerce")
                # UNITS/TALLA(pedido): usar TT si existe; fallback a UNITS/TALLA (PEDIDO)
                if "TT" in df_merge_all.columns:
                    df_merge_all["UNITS/TALLA(pedido)"] = df_merge_all.get("TT", "")
                else:
                    df_merge_all["UNITS/TALLA(pedido)"] = df_merge_all.get("UNITS/TALLA (PEDIDO)", "")
                df_merge_all["WIP Line Number"] = df_merge_all.get("WIP LINE NUMBER", "")
                df_merge_all["STYLE/COLOR"] = df_merge_all.get("STYLE COLOR", "")
                df_merge_all["UPC Barcode"] = df_merge_all.get("UPC CODE", "")

                def make_skx(po: object) -> str:
                    s = str(po or "").strip()
                    if s == "":
                        return ""
                    return s if s.upper().startswith("P") else "P" + s

                df_merge_all["SKX PO#"] = df_merge_all["PO(cliente)"].map(make_skx)

                # Case QTY desde Excel (agregar prefijo Q)
                if "CASE QTY" in df_merge_all.columns:
                    proc.update_status("Extrayendo Case QTY del Excel…")
                    def _q_prefix(v: object) -> str:
                        s = str(v).strip()
                        if s == "" or s.lower() == "nan":
                            return ""
                        # Extraer solo la parte numérica (ej: PP10 -> 10, Q60 -> 60, 75 -> 75)
                        digits = re.sub(r"[^0-9]", "", s)
                        if not digits:
                            return ""
                        return "Q" + digits
                    df_merge_all["Case QTY"] = df_merge_all["CASE QTY"].map(_q_prefix)
                else:
                    df_merge_all["Case QTY"] = ""

                df_merge_all["QTY DE STICKERS A IMPRIMIR"] = ""

                columnas_final = [
                    "PROTO",
                    "OP",
                    "PO(cliente)",
                    "UNITS/TALLA(pedido)",
                    "SKX PO#",
                    "WIP Line Number",
                    "STYLE/COLOR",
                    "UPC Barcode",
                    "Case QTY",
                    "US Size",
                    "QTY POR TALLA",
                    "QTY DE STICKERS A IMPRIMIR",
                ]

                for col in columnas_final:
                    if col not in df_merge_all.columns:
                        df_merge_all[col] = ""

                df_merge_all["SIZE_SORTED"] = pd.Categorical(df_merge_all["US Size"], categories=SIZE_ORDER, ordered=True)
                df_final = df_merge_all.sort_values(by=["SKX PO#", "WIP Line Number", "STYLE/COLOR", "SIZE_SORTED"]).drop(columns=["SIZE_SORTED"]).reset_index(drop=True)

                proc.update_status("Generando archivo final…")

                output_dir = Path(pdf_paths[0]).parent
                template_ext = Path(self.state.template_excel_path).suffix.lower()
                out_name = "reporte_final_case_content" + (".xlsm" if template_ext == ".xlsm" else ".xlsx")
                final_filename = str(output_dir / out_name)

                try:
                    if os.path.exists(final_filename):
                        os.remove(final_filename)
                except Exception:
                    pass

                shutil.copy2(self.state.template_excel_path, final_filename)
                wb = openpyxl.load_workbook(final_filename, keep_vba=template_ext == ".xlsm")

                # Determinar hoja plantilla (primera hoja)
                template_sheet = wb.worksheets[0]
                first_style = True

                for style_name, df_style in df_final.groupby("NOMBRE ESTILO"):
                    ws = template_sheet if first_style else wb.copy_worksheet(template_sheet)
                    first_style = False
                    # Título hoja máx 31 chars
                    ws.title = str(style_name)[:31] if str(style_name).strip() else ws.title

                    # Colocar imagen (reemplaza cualquier imagen previa)
                    copy_template_header_to_worksheet(template_sheet, ws, self.state.img1_path)

                    df_out = df_style[columnas_final].copy()
                    data_start_row = 10  # fila de encabezados

                    # ESCRIBIR ENCABEZADOS en la fila 10
                    for col_idx, col_name in enumerate(columnas_final, 1):
                        cell = ws.cell(row=data_start_row, column=col_idx)
                        cell.value = col_name
                        cell.font = Font(bold=True, color="000000")  # Negro y negrita
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo amarillo
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # AGREGAR AJUSTAR TEXTO

                    # QUITAR LÍNEAS DE CUADRÍCULA DE CADA HOJA
                    ws.sheet_view.showGridLines = False

                    # AGREGAR COLORES A LA FILA 11 (fila vacía entre encabezados y datos)
                    empty_row = data_start_row + 1  # Fila 11
                    
                    # Celda F11 - Verde176 + Azul80 (RGB: 0, 176, 80)
                    ws.cell(row=empty_row, column=6).fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                    
                    # Celdas G11 y H11 (COMBINADAS) - AZUL ENFASIS 1 OSCURO 25% (RGB: 68, 114, 196)
                    ws.merge_cells(f'G{empty_row}:H{empty_row}')
                    ws.cell(row=empty_row, column=7).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    
                    # Celda I11 - Rojo255 + Verde192 (RGB: 255, 192, 0)
                    ws.cell(row=empty_row, column=9).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    
                    # Celda J11 - Rojo112 + Verde48 + Azul160 (RGB: 112, 48, 160)
                    ws.cell(row=empty_row, column=10).fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

                    # Celda L6 - Verde176 + Azul80 (RGB: 0, 176, 80)
                    ws.cell(row=6, column=12).fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

                    # DEJAR FILA 11 VACÍA (pero con colores) - Solo agregar datos a partir de la fila 12
                    first_data_row = data_start_row + 2  # Fila 10 = encabezados, Fila 11 = vacía con colores, Fila 12 = datos

                    # Datos solamente (empezando en fila 12)
                    for row_idx, (_, row_data) in enumerate(df_out.iterrows(), first_data_row):
                        for col_idx, value in enumerate(row_data, 1):
                            if col_idx <= len(columnas_final):  # Solo escribir en las columnas que corresponden
                                cell = ws.cell(row=row_idx, column=col_idx)
                                
                                # CONVERTIR A NÚMERO LAS COLUMNAS QUE DEBEN SER NUMÉRICAS
                                column_name = columnas_final[col_idx - 1]  # -1 porque col_idx empieza en 1
                                
                                # Columnas que deben mantenerse como texto
                                text_columns = ["SKX PO#", "STYLE/COLOR", "Case QTY", "US Size"]
                                
                                if column_name not in text_columns and value:
                                    # Intentar convertir a número
                                    try:
                                        # Limpiar el valor de espacios y caracteres no numéricos innecesarios
                                        clean_value = str(value).strip()
                                        if clean_value and clean_value != 'nan' and clean_value != '':
                                            # Intentar convertir a entero primero, luego a float
                                            if '.' in clean_value:
                                                numeric_value = float(clean_value)
                                                # Si es un entero disfrazado de float, convertir a int
                                                if numeric_value.is_integer():
                                                    cell.value = int(numeric_value)
                                                else:
                                                    cell.value = numeric_value
                                            else:
                                                cell.value = int(clean_value)
                                        else:
                                            cell.value = value
                                    except (ValueError, TypeError):
                                        # Si no se puede convertir, mantener como texto
                                        cell.value = value
                                else:
                                    # Para columnas de texto, mantener como string
                                    cell.value = value
                                
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                                # Mantener contenido sin negrita
                                cell.font = Font(bold=False)

                    # Calcular last_data_row directamente del número de filas escritas
                    last_data_row = first_data_row + len(df_out) - 1

                    # Limpiar cualquier contenido residual de la plantilla debajo de los datos
                    for r in range(last_data_row + 1, ws.max_row + 1):
                        for c in range(1, len(columnas_final) + 1):
                            cell = ws.cell(row=r, column=c)
                            if not isinstance(cell, MergedCell):
                                cell.value = None

                    apply_formulas_to_sheet(ws, data_start_row, len(columnas_final), last_data_row)

                # Si copiamos hojas, elimina la plantilla si es genérica
                if len(wb.worksheets) > 1 and template_sheet.title.lower() in [
                    "plantilla", "template", "sheet1", "hoja1", "encabezado"
                ]:
                    wb.remove(template_sheet)

                wb.save(final_filename)

                # Cerrar ventana y preguntar si abrir
                proc.close()
                def _done():
                    response = messagebox.askyesno(
                        "Proceso Completado",
                        f"✅ Archivo generado exitosamente:\n{final_filename}\n\n¿Desea abrir el archivo ahora?",
                        icon="question",
                    )
                    if response:
                        open_file_and_folder(final_filename)
                self.root.after(0, _done)

            except Exception as e:
                proc.close()
                # Imprimir traza completa en stderr para depuración (temporal)
                try:
                    import traceback, sys
                    traceback.print_exc(file=sys.stderr)
                except Exception:
                    pass
                self.root.after(0, lambda e=e: messagebox.showerror("Error", f"Error durante el procesamiento: {e}"))

        threading.Thread(target=worker, daemon=True).start()

# ==========================
#  main
# ==========================

def main() -> None:
    root = tk.Tk()
    # Estilo ttk simple
    try:
        style = ttk.Style()
        style.theme_use("clam")
    except Exception:
        pass
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
