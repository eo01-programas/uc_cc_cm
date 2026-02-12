import os
import re
import sys
import platform
import subprocess
import unicodedata
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from PIL import Image, ImageTk

import pandas as pd
import pdfplumber
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


# ============================================================
#  GENERADOR DE REPORTE FINAL (PDF UPC + EXCEL RSV/OP)
#  Fixes incluidos:
#   - Acepta "RSV" como reemplazo de "OP" si OP no existe
#   - Acepta "DESCRIPCION DE COLOR" además de "DESCRIPCION COLOR"
#   - Mejoras: filtro automático a todas las columnas
# ============================================================


# ==========================
#  RUTEO DE RECURSOS (VSCode / PyInstaller)
# ==========================

def app_dir() -> Path:
    """Devuelve la carpeta base de la app.
    - En PyInstaller (_MEIPASS) usa el directorio temporal del bundle
    - En modo script usa la carpeta del archivo .py
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent

BASE_DIR = app_dir()


def locate_asset(base_name: str, exts: list[str], extra_dirs: list[Path] | None = None) -> str:
    """Busca `base_name` con cualquiera de `exts` en:
    - carpeta del script (BASE_DIR)
    - carpeta de trabajo (CWD)
    - carpetas adicionales (por ejemplo, donde están los PDFs/Excel)
    Devuelve la ruta encontrada o "" si no existe.
    """
    search_dirs = [BASE_DIR, Path.cwd()]
    if extra_dirs:
        search_dirs.extend(extra_dirs)
    for d in search_dirs:
        for ext in exts:
            p = d / f"{base_name}{ext}"
            if p.exists():
                return str(p)
    return ""


# Valores iniciales (se re-ubican tras elegir archivos)
header_path = locate_asset("encabezado", [".xlsx"])
img1_path   = locate_asset("imagen1", [".png", ".jpg", ".jpeg", ".bmp"])
img2_path   = locate_asset("imagen2", [".png", ".jpg", ".jpeg", ".bmp"])


# ==========================
#  UI: PREVIEW / CAMBIO IMÁGENES
# ==========================

def mostrar_preview(path: str, label: tk.Label) -> None:
    try:
        if not path or not os.path.exists(path):
            raise FileNotFoundError(path or "(vacío)")
        img = Image.open(path)
        img.thumbnail((100, 100))
        photo = ImageTk.PhotoImage(img)
        label.config(image=photo, text="")
        label.image = photo
    except Exception:
        label.config(text="No se pudo cargar", image="")
        label.image = None


def cambiar_imagen(n: int) -> None:
    global img1_path, img2_path
    file = filedialog.askopenfilename(
        title=f"Selecciona la imagen {n}",
        filetypes=[("Imágenes", "*.png;*.jpg;*.jpeg;*.bmp")]
    )
    if file:
        if n == 1:
            img1_path = file
            mostrar_preview(img1_path, lbl_img1)
        else:
            img2_path = file
            mostrar_preview(img2_path, lbl_img2)


# ==========================
#  COPIAR ENCABEZADO (filas 1..13) + IMÁGENES
# ==========================

def copiar_encabezado(ws_origen, ws_destino, filas: int = 13) -> None:
    from copy import copy

    # Copia valores/estilos
    for row in ws_origen.iter_rows(min_row=1, max_row=filas):
        for cell in row:
            if cell.__class__.__name__ == "MergedCell":
                continue

            # Ajustes de celdas especiales según tu plantilla
            if cell.coordinate in ["C2", "C3", "C4"]:
                new_cell = ws_destino.cell(row=cell.row, column=2, value=cell.value)
            elif cell.coordinate in ["B10", "B11", "B12"]:
                new_cell = ws_destino.cell(row=cell.row, column=1, value=cell.value)
            elif cell.coordinate == "I3":
                new_cell = ws_destino.cell(row=3, column=7, value=cell.value)
            elif cell.coordinate == "J6":
                new_cell = ws_destino.cell(row=6, column=9, value=cell.value)
            else:
                new_cell = ws_destino.cell(row=cell.row, column=cell.column, value=cell.value)

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)

    # Copia merges
    for merge in ws_origen.merged_cells.ranges:
        if merge.min_row <= filas:
            ws_destino.merge_cells(str(merge))

    # Inserta imágenes
    try:
        if img1_path and os.path.exists(img1_path):
            img1 = XLImage(img1_path)
            img1.width = 6.5 * 37.7952755906
            img1.height = 6.5 * 37.7952755906
            ws_destino.add_image(img1, "B5")
        if img2_path and os.path.exists(img2_path):
            img2 = XLImage(img2_path)
            img2.width = 7.0 * 37.7952755906
            img2.height = 6.5 * 37.7952755906
            ws_destino.add_image(img2, "F5")
    except Exception as e:
        print(f"No se pudo insertar una imagen: {e}")

    # Anchos/altos base
    col_widths = {
        'A': 20, 'B': 20, 'C': 41, 'D': 14.22, 'E': 14.22, 'F': 9,
        'G': 13.22, 'H': 12.22, 'I': 8.11, 'J': 6.22, 'K': 30
    }
    for col, width in col_widths.items():
        ws_destino.column_dimensions[col].width = width

    for row in range(5, 10):
        ws_destino.row_dimensions[row].height = 40

    # Textos en K
    font_bold_10 = Font(bold=True, size=10)
    alignment_wrap = Alignment(wrap_text=True, horizontal="center", vertical="center")

    ws_destino['K5'].value = "MADE IN PERU FABRIQUÉ AU PÉROU"
    ws_destino['K5'].font = font_bold_10
    ws_destino['K5'].alignment = alignment_wrap
    ws_destino['K5'].value = ws_destino['K5'].value.upper()

    ws_destino.merge_cells('K6:K7')
    ws_destino['K6'].value = "SIZE/TAILLE:"
    ws_destino['K6'].font = font_bold_10
    ws_destino['K6'].alignment = alignment_wrap
    ws_destino['K6'].value = ws_destino['K6'].value.upper()

    ws_destino['K8'].value = "COLOR/COULEUR:"
    ws_destino['K8'].font = font_bold_10
    ws_destino['K8'].alignment = alignment_wrap
    ws_destino['K8'].value = ws_destino['K8'].value.upper()

    # Fuente más grande en K
    for row in ws_destino.iter_rows(min_row=1, max_row=ws_destino.max_row, min_col=11, max_col=11):
        for cell in row:
            cell.font = Font(size=16, bold=cell.font.bold if cell.font else False)


# ==========================
#  PDF: DETECCIÓN Y EXTRACCIÓN ROBUSTA
# ==========================

def detectar_formato(pdf_path: str) -> str:
    try:
        with pdfplumber.open(pdf_path) as doc:
            text = (doc.pages[0].extract_text() or "")
            if "Division|" in text:
                return "Barras"
            if "UPC REPORT" in text:
                return "Matricial"
    except Exception:
        pass
    return "Desconocido"


def extract_data_barras(pdf_path: str) -> list[dict]:
    data: list[dict] = []
    with pdfplumber.open(pdf_path) as doc:
        full_text = "\n".join([page.extract_text() or "" for page in doc.pages])

    lines = [ln.strip() for ln in full_text.split("\n") if ln.strip()]
    for line in lines:
        if 'Division|' in line and 'Style|' in line and 'UPC|' in line:
            continue
        if '|' not in line:
            continue

        parts = [p.strip() for p in line.split('|')]
        if len(parts) < 8:
            continue

        # layout típico: Division|Style|UPC|Style Name|Color Code|Color Name|Size Group|Size
        _, style, upc, _, color_code, color_name, _, size = parts[:8]

        upc_digits = re.sub(r"\D", "", str(upc))
        # UPC suele ser 11-14 dígitos; valida mínimo 11
        if len(upc_digits) < 11:
            continue

        style_u = str(style).strip().upper()
        color_code_u = str(color_code).strip().upper()
        color_name_u = str(color_name).strip().upper()
        size_u = str(size).strip().upper()

        data.append({
            'STYLE': style_u,
            'COLOR CODE': color_code_u,
            'COLOR NAME': color_name_u,
            'SIZE': size_u,
            'UPC CODE': upc_digits,
            'STYLE COLOR': f"{style_u} {color_code_u}",
        })
    return data


# ---- Matricial (UPC REPORT BY STYLE/COLOR) ----
def extract_data_matricial(pdf_path: str) -> list[dict]:
    registros: list[dict] = []
    style_actual: str | None = None
    tallas_actuales: list[str] = []

    # Generaliza: TP214, TS167, etc.
    style_re = re.compile(r"^([A-Z]{2}\d+[A-Z]?)\b")
    size_token = re.compile(r"\*+\s*([A-Z0-9/]+)\s*\*+")
    color_line = re.compile(
        r"^([A-Z0-9]{3,5})\s+([A-Z0-9/ .\-]+?)(?:\s+((?:\d{11,14}\s+)*\d{11,14}))?\s*$"
    )
    numbers_only = re.compile(r"^(?:\d{11,14}\s+)*\d{11,14}$")

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            raw_lines = [ln.rstrip() for ln in text.split("\n")]
            i = 0
            while i < len(raw_lines):
                line = raw_lines[i].strip()
                if not line or line.startswith("-"):
                    i += 1
                    continue

                # 1) Detecta estilo y tallas
                m_style = style_re.match(line)
                if m_style:
                    style_actual = m_style.group(1).upper()
                    tallas_actuales = size_token.findall(line)

                    # tallas en líneas siguientes
                    j = i + 1
                    while j < len(raw_lines):
                        nxt = raw_lines[j].strip()
                        if not nxt:
                            break
                        if style_re.match(nxt) or color_line.match(nxt):
                            break
                        extra_sizes = size_token.findall(nxt)
                        if not extra_sizes:
                            break
                        tallas_actuales.extend(extra_sizes)
                        j += 1

                    i = j
                    continue

                # 2) Detecta color y UPCs (con continuaciones)
                m_color = color_line.match(line)
                if m_color and style_actual and tallas_actuales:
                    color_code = m_color.group(1).upper()
                    color_name = m_color.group(2).strip().upper()
                    upcs: list[str] = []
                    if m_color.group(3):
                        upcs.extend(m_color.group(3).split())

                    k = i + 1
                    while k < len(raw_lines):
                        nxt = raw_lines[k].strip()
                        if numbers_only.match(nxt):
                            upcs.extend(nxt.split())
                            k += 1
                            continue
                        if color_line.match(nxt) or style_re.match(nxt):
                            break
                        if not nxt:
                            break
                        break

                    n = min(len(tallas_actuales), len(upcs))
                    for idx in range(n):
                        size = tallas_actuales[idx].upper()
                        registros.append({
                            "STYLE": style_actual,
                            "COLOR CODE": color_code,
                            "COLOR NAME": color_name,
                            "SIZE": size,
                            "UPC CODE": upcs[idx],
                            "STYLE COLOR": f"{style_actual} {color_code}",
                        })

                    i = k
                    continue

                i += 1

    return registros


# ==========================
#  NORMALIZACIÓN TALLAS + EXCEL
# ==========================

SIZE_MAP = {
    'XXS': 'XXS', 'XS': 'XS', 'S': 'S', 'M': 'M', 'L': 'L', 'XL': 'XL',
    'XXL': '2XL', '2XL': '2XL', 'XXXL': '3XL', '3XL': '3XL',
    'XSS': 'XS', 'SMALL': 'S', 'MEDIUM': 'M', 'LARGE': 'L',
    'EXTRA SMALL': 'XS', 'EXTRA LARGE': 'XL',
    'EXTRA EXTRA LARGE': '2XL', 'EXTRA EXTRA EXTRA LARGE': '3XL'
}
SIZE_ORDER = ['XXS', 'XS', 'S', 'M', 'L', 'XL', '2XL', '3XL']
SIZE_CANONICAL = set(SIZE_ORDER)

CANADA_SIZE_MAP = {
    'S': 'S/P',
    'M': 'M/M',
    'L': 'L/G',
    'XL': 'XL/TG',
    '2XL': '2XL/TTG',
    '3XL': '3XL/TTTG',
}

BRAZIL_SIZE_MAP = {
    'XS': 'XS/PP',
    'S': 'S/P',
    'M': 'M/M',
    'L': 'L/G',
    'XL': 'XL/GG',
    '2XL': 'XXL/XGG',
    'XXL': 'XXL/XGG',
}

def norm_size(s):
    if s is None:
        return s
    su = str(s).strip().upper()
    return SIZE_MAP.get(su, su)


def leer_excel_flexible(excel_path: str) -> pd.DataFrame:
    """Detecta la fila de encabezado buscando tokens típicos.
    Si no detecta, cae al fallback.
    """
    def norm_token(value: str) -> str:
        raw = unicodedata.normalize("NFKD", str(value).strip())
        raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
        cleaned = re.sub(r"[^A-Z0-9#]+", " ", raw.upper())
        return re.sub(r"\s+", " ", cleaned).strip()

    tokens = {
        "STYLE", "ESTILOS", "OP", "RSV", "PROTO", "DESTINO", "PO", "PO NO", "PO NO.", "PO#",
        "DESCRIPCION COLOR", "DESCRIPCION DE COLOR", "COLOR", "CARTA", "CODE", "COLR CODE", "COLOR CODE",
        "LN"
    }

    best = None  # (hits_count, sheet, header_row)

    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
        for sheet in xls.sheet_names:
            preview = pd.read_excel(excel_path, sheet_name=sheet, header=None, nrows=40, engine="openpyxl")
            for idx, row in preview.iterrows():
                cells = [norm_token(c) for c in row.tolist()]
                hits = {c for c in cells if c in tokens}
                if ("STYLE" in hits) or ("ESTILOS" in hits):
                    hits_count = len(hits)
                    if best is None or hits_count > best[0]:
                        best = (hits_count, sheet, idx)
    except Exception:
        best = None

    if best:
        _, sheet, header_row = best
        return pd.read_excel(excel_path, dtype=str, engine="openpyxl", sheet_name=sheet, header=header_row)

    # Fallback conservador
    return pd.read_excel(excel_path, dtype=str, engine="openpyxl", sheet_name=0, header=1)


def preparar_excel(df_excel: pd.DataFrame) -> pd.DataFrame:
    """Renombra columnas a nombres internos y (si hay tallas como columnas)
    deja un registro por talla (columna SIZE).
    """
    def norm_col(name: str) -> str:
        raw = unicodedata.normalize("NFKD", str(name).strip())
        raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
        cleaned = re.sub(r"[^A-Z0-9#]+", " ", raw.upper())
        return re.sub(r"\s+", " ", cleaned).strip()

    cols_norm = {norm_col(c): c for c in df_excel.columns}

    def pick_col(*candidates: str) -> str | None:
        for cand in candidates:
            if cand in cols_norm:
                return cols_norm[cand]
        return None

    rename_map: dict[str, str] = {}

    style_col = pick_col('ESTILOS', 'STYLE')
    if style_col:
        rename_map[style_col] = 'NOMBRE ESTILO'

    # OP o RSV (fix principal)
    op_col = pick_col('OP')
    rsv_col = pick_col('RSV')
    if op_col:
        rename_map[op_col] = 'PEDIDO PRODUCCION COFACO'
    elif rsv_col:
        rename_map[rsv_col] = 'PEDIDO PRODUCCION COFACO'

    proto_col = pick_col('PROTO')
    if proto_col:
        rename_map[proto_col] = 'PROTO COFACO'

    destino_col = pick_col('DESTINO')
    if destino_col:
        rename_map[destino_col] = 'DESTINO'

    po_col = pick_col('PO', 'PO NO', 'PO NO.', 'PO#')
    if po_col:
        rename_map[po_col] = 'PO#'

    # (Opcional) LN si existe
    ln_col = pick_col('LN')
    if ln_col:
        rename_map[ln_col] = 'LN'

    # Nombre de color (fix: incluye "DESCRIPCION DE COLOR")
    nombre_color_col = pick_col('DESCRIPCION COLOR', 'DESCRIPCION DE COLOR')
    if not nombre_color_col:
        nombre_color_col = pick_col('COLOR', 'CARTA')
    if nombre_color_col:
        rename_map[nombre_color_col] = 'NOMBRE COLOR'

    # Código de color
    codigo_color_col = pick_col('COLR CODE', 'COLOR CODE')
    if not codigo_color_col:
        codigo_color_col = pick_col('CODE')
    if not codigo_color_col:
        color_col = cols_norm.get('COLOR')
        if color_col and color_col != nombre_color_col:
            codigo_color_col = color_col
    if codigo_color_col:
        rename_map[codigo_color_col] = 'COLOR'

    df_excel = df_excel.rename(columns=rename_map)

    # Requeridas (OP ya no es obligatorio si hay RSV)
    requeridas = ['NOMBRE ESTILO', 'PEDIDO PRODUCCION COFACO', 'PROTO COFACO', 'DESTINO', 'PO#', 'NOMBRE COLOR']
    falt = [c for c in requeridas if c not in df_excel.columns]
    if falt:
        raise ValueError(f"Faltan columnas en el Excel: {falt}")

    if 'COLOR' not in df_excel.columns:
        df_excel['COLOR'] = ""

    # Normaliza texto
    for c in ['NOMBRE ESTILO', 'DESTINO', 'NOMBRE COLOR', 'COLOR']:
        df_excel[c] = df_excel[c].astype(str).str.strip().str.upper()

    # Detecta tallas como columnas (S, M, L, XL, 2XL, 3XL, etc.)
    size_cols = []
    for c in df_excel.columns:
        cu = str(c).strip().upper()
        if cu in SIZE_CANONICAL or cu in SIZE_MAP:
            size_cols.append(c)

    if size_cols:
        canon = {c: SIZE_MAP.get(str(c).strip().upper(), str(c).strip().upper()) for c in size_cols}

        id_vars = ['NOMBRE ESTILO', 'PEDIDO PRODUCCION COFACO', 'PROTO COFACO', 'DESTINO', 'PO#', 'NOMBRE COLOR', 'COLOR']
        if 'LN' in df_excel.columns:
            id_vars.append('LN')

        df_long = df_excel.melt(
            id_vars=id_vars,
            value_vars=size_cols,
            var_name='SIZE_RAW',
            value_name='QTY'
        )
        df_long['SIZE'] = df_long['SIZE_RAW'].map(lambda x: canon.get(x, str(x).upper()))

        def has_qty(v) -> bool:
            if v is None:
                return False
            s = str(v).strip()
            if s in ("", "0", "0.0"):
                return False
            try:
                return float(s) > 0
            except Exception:
                return True

        df_long = df_long[df_long['QTY'].apply(has_qty)].copy()
        df_long['SIZE'] = df_long['SIZE'].map(norm_size)

        df_long['SIZE_SORTED'] = pd.Categorical(df_long['SIZE'], categories=SIZE_ORDER, ordered=True)
        sort_cols = ['NOMBRE ESTILO', 'DESTINO', 'NOMBRE COLOR', 'SIZE_SORTED']
        df_long = df_long.sort_values(sort_cols).drop(columns=['SIZE_SORTED'])

        keep = id_vars + ['SIZE']
        return df_long[keep].reset_index(drop=True)

    # Si no hay columnas de tallas, devuelve estructura única (sin SIZE)
    base_cols = ['NOMBRE ESTILO', 'PEDIDO PRODUCCION COFACO', 'PROTO COFACO', 'DESTINO', 'PO#', 'NOMBRE COLOR', 'COLOR']
    if 'LN' in df_excel.columns:
        base_cols.append('LN')

    return df_excel[base_cols].drop_duplicates().reset_index(drop=True)


# ==========================
#  PROCESAMIENTO PRINCIPAL
# ==========================

def process_all() -> None:
    global header_path, img1_path, img2_path

    pdf_paths = filedialog.askopenfilenames(
        title="Selecciona los archivos PDF",
        filetypes=[("Archivos PDF", "*.pdf")]
    )
    if not pdf_paths:
        return

    excel_path = filedialog.askopenfilename(
        title="Selecciona el archivo Excel (datos)",
        filetypes=[("Archivos Excel", "*.xlsx;*.xls;*.xlsm")]
    )
    if not excel_path:
        return

    status_var.set("Cargando, por favor...")
    root.update_idletasks()

    # Relocaliza recursos
    extra_dirs = [Path(pdf_paths[0]).parent, Path(excel_path).parent]
    if not header_path or not os.path.exists(header_path):
        header_path = locate_asset("encabezado", [".xlsx"], extra_dirs)
    if not img1_path or not os.path.exists(img1_path):
        img1_path = locate_asset("imagen1", [".png", ".jpg", ".jpeg", ".bmp"], extra_dirs)
    if not img2_path or not os.path.exists(img2_path):
        img2_path = locate_asset("imagen2", [".png", ".jpg", ".jpeg", ".bmp"], extra_dirs)

    if not header_path or not os.path.exists(header_path):
        messagebox.showerror(
            "Error",
            "No se encontró 'encabezado.xlsx'. Ponlo junto al .py o en la carpeta de los PDFs/Excel."
        )
        status_var.set("")
        root.update_idletasks()
        return

    # Previews actualizados
    mostrar_preview(img1_path, lbl_img1)
    mostrar_preview(img2_path, lbl_img2)

    # 1) Extrae PDFs
    all_registros: list[dict] = []
    for pdf in pdf_paths:
        tipo = detectar_formato(pdf)
        if tipo == "Barras":
            all_registros.extend(extract_data_barras(pdf))
        else:
            rows = extract_data_matricial(pdf)
            if not rows and tipo == "Desconocido":
                rows = extract_data_barras(pdf)
            all_registros.extend(rows)

    if not all_registros:
        messagebox.showerror("Error", "No se extrajo información de los PDFs.")
        status_var.set("")
        root.update_idletasks()
        return

    df_pdfs = pd.DataFrame(all_registros)
    for c in ['STYLE', 'COLOR CODE', 'COLOR NAME', 'SIZE']:
        if c in df_pdfs.columns:
            df_pdfs[c] = df_pdfs[c].astype(str).str.strip().str.upper()
    df_pdfs['SIZE'] = df_pdfs['SIZE'].map(norm_size)

    # 2) Lee y prepara Excel
    try:
        df_excel_raw = leer_excel_flexible(excel_path)
        df_excel = preparar_excel(df_excel_raw)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo preparar el Excel:\n{e}")
        status_var.set("")
        root.update_idletasks()
        return

    for c in ['NOMBRE ESTILO', 'DESTINO', 'NOMBRE COLOR', 'COLOR']:
        if c in df_excel.columns:
            df_excel[c] = df_excel[c].astype(str).str.strip().str.upper()
    if 'SIZE' in df_excel.columns:
        df_excel['SIZE'] = df_excel['SIZE'].map(norm_size)

    excel_has_sizes = 'SIZE' in df_excel.columns

    # 3) Merge por nombre color (preferido) y fallback por código color
    if excel_has_sizes:
        df_name = pd.merge(
            df_excel, df_pdfs,
            left_on=['NOMBRE ESTILO', 'NOMBRE COLOR', 'SIZE'],
            right_on=['STYLE', 'COLOR NAME', 'SIZE'],
            how='inner'
        )
        df_code = pd.merge(
            df_excel, df_pdfs,
            left_on=['NOMBRE ESTILO', 'COLOR', 'SIZE'],
            right_on=['STYLE', 'COLOR CODE', 'SIZE'],
            how='inner'
        )
    else:
        df_name = pd.merge(
            df_excel, df_pdfs,
            left_on=['NOMBRE ESTILO', 'NOMBRE COLOR'],
            right_on=['STYLE', 'COLOR NAME'],
            how='inner'
        )
        df_code = pd.merge(
            df_excel, df_pdfs,
            left_on=['NOMBRE ESTILO', 'COLOR'],
            right_on=['STYLE', 'COLOR CODE'],
            how='inner'
        )

    df_merge_all = pd.concat([df_name, df_code], ignore_index=True)

    # Dedupe
    dedup_keys = ['NOMBRE ESTILO', 'NOMBRE COLOR', 'COLOR', 'DESTINO', 'PO#', 'UPC CODE']
    if 'SIZE' in df_merge_all.columns:
        dedup_keys.append('SIZE')
    df_merge_all = df_merge_all.drop_duplicates(subset=[k for k in dedup_keys if k in df_merge_all.columns])

    if df_merge_all.empty:
        messagebox.showwarning(
            "Sin resultados",
            "No hubo intersección entre PDFs y Excel con los criterios dados.\n"
            "Tip: revisa que STYLE/ColorCode/Size del PDF coincidan con el Excel."
        )
        status_var.set("")
        root.update_idletasks()
        return

    # 4) Opciones salida
    if jap_var.get():
        df_merge_all['UPC CODE'] = df_merge_all['UPC CODE'].apply(
            lambda s: str(s) if str(s).startswith('0') else '0' + str(s)
        )

    # 5) Selección y orden final
    columnas_final = [
        'PROTO COFACO',
        'PEDIDO PRODUCCION COFACO',
        'DESTINO',
        'NOMBRE ESTILO',
        'NOMBRE COLOR',
        'PO#',
        'UPC CODE',
        'STYLE COLOR',
        'SIZE',
        'COLOR',
        'LN',  # opcional
    ]

    for col in columnas_final:
        if col not in df_merge_all.columns:
            df_merge_all[col] = ""

    df_final = df_merge_all[columnas_final].copy()

    # Orden por tallas antes de Canadá
    if 'SIZE' in df_final.columns:
        df_final['SIZE_SORTED'] = pd.Categorical(df_final['SIZE'], categories=SIZE_ORDER, ordered=True)
        df_final = df_final.sort_values(by=['PEDIDO PRODUCCION COFACO', 'DESTINO', 'PO#', 'NOMBRE COLOR', 'SIZE_SORTED'])
        df_final = df_final.drop(columns=['SIZE_SORTED'])
    else:
        df_final = df_final.sort_values(by=['PEDIDO PRODUCCION COFACO', 'DESTINO', 'PO#', 'NOMBRE COLOR'])

    # Formato Canadá (después del ordenamiento)
    if can_var.get() and 'SIZE' in df_final.columns:
        df_final['SIZE'] = df_final['SIZE'].apply(lambda s: CANADA_SIZE_MAP.get(str(s).upper().strip(), s))

    if br_var.get() and 'SIZE' in df_final.columns:
        df_final['SIZE'] = df_final['SIZE'].apply(lambda s: BRAZIL_SIZE_MAP.get(str(s).upper().strip(), s))

    # 6) Salida
    output_dir = Path(pdf_paths[0]).parent
    name_parts = ["Reporte_Final"]
    if jap_var.get():
        name_parts.append("JP")
    if can_var.get():
        name_parts.append("CA")
    if br_var.get():
        name_parts.append("BR")
    final_filename = str(output_dir / ("_".join(name_parts) + ".xlsx"))

    try:
        if os.path.exists(final_filename):
            os.remove(final_filename)
    except Exception:
        pass

    with pd.ExcelWriter(final_filename, engine="openpyxl") as writer:
        # si por algún motivo está vacío NOMBRE ESTILO, evitar fallo
        if 'NOMBRE ESTILO' not in df_final.columns or df_final['NOMBRE ESTILO'].astype(str).str.strip().eq("").all():
            df_final.to_excel(writer, sheet_name="REPORTE", index=False, startrow=13)
        else:
            for style, df_style in df_final.groupby("NOMBRE ESTILO"):
                sheet = str(style)[:31] if str(style).strip() else "REPORTE"
                df_style.to_excel(writer, sheet_name=sheet, index=False, startrow=13)

    # 7) Aplicar encabezado + formato
    try:
        wb_template = load_workbook(header_path)
        ws_template = wb_template.active
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir '{header_path}':\n{e}")
        status_var.set("")
        root.update_idletasks()
        return

    wb = openpyxl.load_workbook(final_filename)
    for ws in wb.worksheets:
        copiar_encabezado(ws_template, ws, filas=13)

        header_row = 14
        max_row = ws.max_row
        max_col = ws.max_column
        last_col_letter = get_column_letter(max_col)

        # Autofiltro a todas las columnas
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{max_row}"

        # Encabezados tabla
        for cell in ws[header_row]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.font = Font(bold=True)

        # Freeze panes (opcional, útil)
        ws.freeze_panes = f"A{header_row+1}"

        # Autoancho (tope 60), asegurar A=20 y K=30 si existe
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(header_row, max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        ws.column_dimensions['A'].width = 20
        if 'K' in ws.column_dimensions:
            ws.column_dimensions['K'].width = 30

        # Forzar texto
        for row in ws.iter_rows(min_row=header_row + 1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if cell.value is not None:
                    cell.value = str(cell.value).lstrip("'")
                    cell.data_type = "s"

    try:
        wb.save(final_filename)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el Excel final:\n{e}")
        status_var.set("")
        root.update_idletasks()
        return

    try:
        messagebox.showinfo("Éxito", f"Se generó el archivo:\n{final_filename}")
    except Exception:
        pass

    # Abrir archivo y carpeta
    sistema = platform.system()
    try:
        if sistema == "Windows":
            os.startfile(final_filename)
            os.startfile(os.path.dirname(final_filename))
        elif sistema == "Darwin":
            subprocess.Popen(["open", final_filename])
            subprocess.Popen(["open", os.path.dirname(final_filename)])
        else:
            subprocess.Popen(["xdg-open", final_filename])
            subprocess.Popen(["xdg-open", os.path.dirname(final_filename)])
    except Exception:
        pass

    status_var.set("")
    root.update_idletasks()


# ==========================
#  INTERFAZ
# ==========================

root = tk.Tk()
root.title("Generador de Reporte Final")
root.geometry("700x560")

label = tk.Label(
    root,
    text=(
        "Selecciona los PDFs y el Excel de datos.\n"
        "Encabezado (1..13) e imágenes se detectan automáticamente en:\n"
        f"  - {BASE_DIR}\n  - Carpeta actual\n  - Carpeta de los PDFs/Excel.\n"
        "Archivos: encabezado.xlsx, imagen1.(png/jpg), imagen2.(png/jpg)"
    ),
    wraplength=660,
    justify="left"
)
label.pack(pady=10)

status_var = tk.StringVar(value="")
status_label = tk.Label(root, textvariable=status_var, fg="#006400")
status_label.pack(pady=4)

jap_var = tk.BooleanVar(value=False)
can_var = tk.BooleanVar(value=False)
br_var = tk.BooleanVar(value=False)

frame_opts = tk.Frame(root)
frame_opts.pack(pady=5)

chk_japan = tk.Checkbutton(frame_opts, text="Si es para Japón, anteponer '0' al UPC", variable=jap_var)
chk_japan.grid(row=0, column=0, sticky="w", padx=5)

chk_can = tk.Checkbutton(
    frame_opts,
    text="Formato talla Canadá (S/P, M/M, L/G, XL/TG, 2XL/TTG, 3XL/TTTG)",
    variable=can_var
)
chk_can.grid(row=1, column=0, sticky="w", padx=5)

chk_br = tk.Checkbutton(
    frame_opts,
    text="Formato talla Brasil (XS/PP, S/P, M/M, L/G, XL/GG, XXL/XGG)",
    variable=br_var
)
chk_br.grid(row=2, column=0, sticky="w", padx=5)

frm_imgs = tk.Frame(root)
frm_imgs.pack(pady=10)

lbl_img1 = tk.Label(frm_imgs, text="Imagen 1")
lbl_img1.grid(row=0, column=0, padx=10)
mostrar_preview(img1_path, lbl_img1)

lbl_img2 = tk.Label(frm_imgs, text="Imagen 2")
lbl_img2.grid(row=0, column=1, padx=10)
mostrar_preview(img2_path, lbl_img2)

btn_img1 = tk.Button(frm_imgs, text="Cambiar Imagen 1", command=lambda: cambiar_imagen(1))
btn_img1.grid(row=1, column=0, pady=5)

btn_img2 = tk.Button(frm_imgs, text="Cambiar Imagen 2", command=lambda: cambiar_imagen(2))
btn_img2.grid(row=1, column=1, pady=5)

btn = tk.Button(root, text="Procesar Archivos", command=process_all, height=2, width=30)
btn.pack(pady=20)

root.mainloop()
